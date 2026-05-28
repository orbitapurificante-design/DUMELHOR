#!/usr/bin/env python3
"""
DUMELHOR — Conversor Excel → produtos.json
==========================================
Como usar:
    1. Instalar dependências:  pip install openpyxl
    2. Correr:                 python converter_excel.py SEU_FICHEIRO.xlsx
    3. O ficheiro produtos.json é criado na mesma pasta

Formato esperado do Excel (colunas — order não importa, nomes sim):
    Código   → id do produto
    Nome     → nome do produto
    Preço    → preço (número, ex: 1.25)
    Unidade  → unidade (ex: UN, CX, KG)
    Família  → categoria / família
    Marca    → marca
    Segmento L / M / N  → segmentos opcionais

Nota: Os cabeçalhos são case-insensitive e ignoram acentos.
"""

import sys
import json
import re
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Falta instalar: pip install openpyxl")
    sys.exit(1)


def normalizar(texto):
    """Remove acentos e converte para minúsculas para comparação."""
    if not texto:
        return ''
    nfkd = unicodedata.normalize('NFKD', str(texto))
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


# Mapeamento de possíveis nomes de colunas → campo interno
MAPA_COLUNAS = {
    # Código
    'codigo':    'id',
    'code':      'id',
    'ref':       'id',
    'referencia':'id',
    'ean':       'ean',
    # Nome
    'nome':      'nome',
    'name':      'nome',
    'descricao': 'nome',
    'descr':     'nome',
    'produto':   'nome',
    # Preço
    'preco':     'preco',
    'price':     'preco',
    'pvp':       'preco',
    'valor':     'preco',
    # Unidade
    'unidade':   'un',
    'un':        'un',
    'unit':      'un',
    'uni':       'un',
    # Família / Categoria
    'familia':   'familia',
    'family':    'familia',
    'categoria': 'familia',
    'category':  'familia',
    'cat':       'familia',
    # Marca
    'marca':     'marca',
    'brand':     'marca',
    'fabricante':'marca',
    # Segmentos
    'segmento l':'colL',
    'col l':     'colL',
    'coll':      'colL',
    'segmento m':'colM',
    'col m':     'colM',
    'colm':      'colM',
    'segmento n':'colN',
    'col n':     'colN',
    'coln':      'colN',
}


def converter(ficheiro_excel, saida='produtos.json'):
    wb = openpyxl.load_workbook(ficheiro_excel, data_only=True)
    ws = wb.active

    # Detectar cabeçalhos na primeira linha
    cabecalhos_raw = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    mapa = {}  # índice coluna (1-based) → campo interno
    for i, cab in enumerate(cabecalhos_raw, 1):
        if cab is None:
            continue
        chave = normalizar(str(cab))
        if chave in MAPA_COLUNAS:
            mapa[i] = MAPA_COLUNAS[chave]

    if not mapa:
        print("❌ Nenhuma coluna reconhecida. Verifique os cabeçalhos.")
        print("   Colunas encontradas:", cabecalhos_raw)
        sys.exit(1)

    print(f"✅ Colunas mapeadas: {list(mapa.values())}")

    produtos = []
    for row_idx in range(2, ws.max_row + 1):
        linha = {}
        for col_idx, campo in mapa.items():
            val = ws.cell(row_idx, col_idx).value
            if val is None:
                linha[campo] = ''
            else:
                linha[campo] = val

        # Ignorar linhas sem nome nem código
        if not linha.get('nome') and not linha.get('id'):
            continue

        # Formatar preço
        if 'preco' in linha:
            try:
                linha['preco'] = round(float(str(linha['preco']).replace(',', '.')), 4)
            except (ValueError, TypeError):
                linha['preco'] = 0.0

        # Garantir que id é string
        if 'id' in linha and linha['id']:
            linha['id'] = str(linha['id']).strip()

        # Limpar strings
        for k in ['nome', 'un', 'familia', 'marca', 'colL', 'colM', 'colN']:
            if k in linha and isinstance(linha[k], str):
                linha[k] = linha[k].strip()

        produtos.append(linha)

    # Guardar JSON
    with open(saida, 'w', encoding='utf-8') as f:
        json.dump(produtos, f, ensure_ascii=False, indent=2)

    print(f"✅ {len(produtos)} produtos exportados para {saida}")
    return len(produtos)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python converter_excel.py FICHEIRO.xlsx [saida.json]")
        print("\nExemplo:")
        print("  python converter_excel.py lista_produtos.xlsx")
        print("  python converter_excel.py lista_produtos.xlsx produtos.json")
        sys.exit(0)

    excel = sys.argv[1]
    saida = sys.argv[2] if len(sys.argv) > 2 else 'produtos.json'

    if not Path(excel).exists():
        print(f"❌ Ficheiro não encontrado: {excel}")
        sys.exit(1)

    converter(excel, saida)
